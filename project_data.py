from openpyxl import load_workbook


class TrendData(object):

    PROCESS_AREAS = [{'PP': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'IPM': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'PMC': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'RSKM': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'REQM': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'RD': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'TS': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'PI': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'VER': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'VAL': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'CM': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'MA': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'PPQA': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'DAR': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'SAM': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0},
                     {'ALL PA\'S': 0, 'li': 0, 'pi': 0, 'ni': 0, 'obv': 0}]

    def __init__(self, workbook, worksheet):
        self.workbook = load_workbook(workbook)
        self.worksheet = self.workbook.get_sheet_by_name(worksheet)
        self.get_process_areas()

    def get_process_areas(self):
        pa_col = 0
        rating_col = 0
        # find columns
        for col in range(1, self.worksheet.max_column):
            header = self.worksheet.cell(row=1, column=col).value.strip().lower()
            if header in 'process areas':
                pa_col = col
            if header in 'rating':
                rating_col = col

        # tally up process areas
        for row in range(2, self.worksheet.max_row + 1):
            pa_cell = self.worksheet.cell(row=row, column=pa_col).value.upper()
            rating_cell = self.worksheet.cell(row=row, column=rating_col).value.lower()
            pa = next((item for item in self.PROCESS_AREAS if pa_cell in item), None)
            try:
                pa[rating_cell] += 1
                pa['ALL PA\'S'] += 1
            except KeyError:
                pass

""" Write to spreadsheet. """

from openpyxl.styles import Alignment, PatternFill, Border, Side


class WriteData(object):
    """ Write to spreadsheet. """

    COLOR_CODE = {'LI': '92D050',
                  'PI': 'FFFF00',
                  'OBV': 'FFC000',
                  'NI': 'FF0000'}

    def __init__(self, worksheet, table_data, project_info):
        self.worksheet = worksheet
        self.table_data = table_data
        self.project_info = project_info
        self.row_offset = 0
        self.col_offset = 0
        self.process_areas = {}
        self.project_info.update({'#Findings': 0})

    def write_to_sheet(self):
        """ Write extracted data to spreadsheet. """
        self.get_offsets()
        for row in range(len(self.table_data)):
            self.write_table_data(row)
            self.write_project_info(row)

    def get_offsets(self):
        """ Get row and column offsets. """
        for row in self.worksheet.iter_rows():
            for cell in row:
                # first empty row
                self.row_offset = cell.row + 1
                # starting cell of findings table
                if cell.value == 'Process Area':
                    self.col_offset = ord(cell.column)
                    break

    def write_table_data(self, row):
        """ Write data in table to sheet. """
        for col in range(self.col_offset, self.col_offset + 5):
            new_row = row + self.row_offset
            new_col = col - 64
            header_info = self.worksheet.cell(row=1, column=new_col).value
            working_cell = self.worksheet.cell(row=new_row, column=new_col)
            align = 'center'

            # put data into cell
            working_cell.value = str(self.table_data[row][new_col - 5])

            if header_info == 'Finding':
                align = 'general'
            elif header_info == 'Rating':
                # update rating value
                if working_cell.value not in self.project_info:
                    self.project_info.update({working_cell.value: 1})
                else:
                    self.project_info[working_cell.value] += 1
                self.project_info['#Findings'] += 1

                color = self.pick_rating_color(working_cell.value.upper())
                working_cell.fill = PatternFill(fill_type='solid',
                                                start_color=color)
                working_cell.border = Border(left=Side(border_style='thin'),
                                             right=Side(border_style='thin'),
                                             top=Side(border_style='thin'),
                                             bottom=Side(border_style='thin'))

            working_cell.alignment = Alignment(horizontal=align,
                                               vertical='center',
                                               wrap_text=True)

    def write_project_info(self, row):
        """ Write project information to sheet. """
        for col in range(1, 5, 1):
            # Styling has to be applied directly to working cell each time
            working_cell = self.worksheet.cell(row=self.row_offset + row, column=col)
            header_info = self.worksheet.cell(row=1, column=col).value.strip(' ')
            try:
                working_cell.value = self.project_info[header_info]
            except KeyError:
                pass
            working_cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=True)

    def pick_rating_color(self, value):
        """ Pick the fill color for the "Rating" field. """
        value = value.strip()
        if value in self.COLOR_CODE:
            return self.COLOR_CODE[value]
        else:
            return 'FFFFFF'

    # might not be as viable as previously thought
    # def get_process_areas(self):
    #     pa_col = 0
    #     # Look for process areas
    #     for col in range(1, self.worksheet.max_column):
    #         header_info = self.worksheet.cell(row=1, column=col).value.lower().strip(' ')
    #         if header_info in 'process areas':
    #             pa_col = col
    #             break
    #
    #     for row in range(2, self.worksheet.max_row + 1):
    #         working_cell = self.worksheet.cell(row=row, column=pa_col).value.upper()
    #         if working_cell not in self.process_areas:
    #             self.process_areas.update({working_cell: 1})
    #         else:
    #             self.process_areas[working_cell] += 1
    #
    #     return self.process_areas

    def get_project_info(self):
        return self.project_info

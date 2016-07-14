""" Extract data from a specific table in Final CAPA reports
    and put it into trending analysis spreadsheet. """

# import sys
# import re
import get_data
import write_data
from openpyxl.styles import Alignment, PatternFill, Border, Side

""" Some values will have to be hardcoded in. This is because there
    are some inconsistencies in the Final CAPA's.

    Inconsistent:
    - Two leads for one project, separated by "/" E.g. "Adam/Monika"
    - Batch name sometimes not included under "Project Information"
    - Customer name is sometimes the site name, other times it is
      site name + entity
    - Different formatting/spelling of go live date E.g. "GO date" or
      "Go live date" or "Go-Live date"
    - Detail of Findings table is sometimes 2nd from last table or 3rd
      from last
        - some reports have extra color code table at the end

    Consistent:
    - Batch name is always 3rd non-empty line in document
    - followed by lead's name
    - followed by report date
    - SAP ID
    - "Project Stakeholders" section follows the same order
    - Detail of Findings table have consistent header
        - Process Area   Goal   Practice   Descrition   Rating """

COLOR_CODE = {'LI':'92D050',
              'PI':'FFC000',
              'Obv':'FFC000',
              'NI':'FF0000'}



def main():
    """ Run the program. """
    print('Loading...')
    # get data from doc
    project = get_data.GetData('TestSheet.xlsx', 'June Data', 'Example_Doc.docx')
    project.process_document()
    worksheet = project.get_worksheet()
    workbook = project.get_workbook()
    table_data = project.get_table_data()
    project_info = project.get_project_info()

    # write to spreadsheet
    do_write = write_data.WriteData(worksheet, table_data, project_info)
    do_write.write_to_sheet()

    # save changes made
    workbook.save('TestSheet.xlsx')
    print('Done!')

def get_offsets(worksheet):
    """ Get the offset to start printing table data into spreadsheet. """
    offsets = [0] * 2
    for row in worksheet.iter_rows():
        for cell in row:
            # starting cell of findings table
            if cell.value == 'Process Area':
                offsets[0] = cell.column
                break
            elif cell.value is None:
                offsets[1] = cell.row + 1
                break

    return offsets

def fill_in_table_data(worksheet, table_data, row, row_offset, col_offset):
    """ Put findings data into the spreadsheet. """
    for col in range(col_offset, col_offset + 5):
        new_row = row + row_offset
        new_col = col - 64
        header_info = worksheet.cell(row=1, column=new_col).value
        working_cell = worksheet.cell(row=new_row, column=new_col)
        align = 'center'
        # put data into cell
        working_cell.value = str(table_data[row][new_col - 5])

        if header_info == 'Finding':
            align = 'general'
        elif header_info == 'Rating':
            color = pick_rating_color(working_cell.value)
            if color == '':
                raise ValueError('Rating value is not valid.')

            working_cell.fill = PatternFill(fill_type='solid',
                                            start_color=color)
            working_cell.border = Border(left=Side(border_style='thin'),
                                         right=Side(border_style='thin'),
                                         top=Side(border_style='thin'),
                                         bottom=Side(border_style='thin'))

        working_cell.alignment = Alignment(horizontal=align,
                                           vertical='center',
                                           wrap_text=True)

def pick_rating_color(value):
    """ Pick the fill color for the "Rating" field. """
    return COLOR_CODE[value]

def fill_in_project_info(worksheet, project_info, row, row_offset):
    """ Put the project's information into the spreadsheet. """
    for col2 in range(1, 5, 1):
        working_cell = worksheet.cell(row=row_offset + row, column=col2)
        header_info = worksheet.cell(row=1, column=col2).value.strip(' ')

        try:
            working_cell.value = project_info[header_info]
        except ValueError:
            print('Header info does not match what\'s in dictionary.')
            print('worksheet: {0: >20}'.format(worksheet))
            print('project_info: {0: >20}'.format(project_info[header_info]))
            print('row: {0: >20}'.format(row))
            print('row_offset: {0: >20}'.format(row_offset))

        working_cell.alignment = Alignment(horizontal='center',
                                           vertical='center',
                                           wrap_text=True)

if __name__ == '__main__':
    main()
